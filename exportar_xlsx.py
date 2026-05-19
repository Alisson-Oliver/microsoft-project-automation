from pathlib import Path
from datetime import datetime

import win32com.client
import pythoncom
import openpyxl

from config import PROJECT_VISIVEL, COLUNAS, PASTA_XLSX, SUFIXO_COPIA

_CAMPOS_DATA = {"Start", "Finish", "ActualStart", "ActualFinish",
                "BaselineStart", "BaselineFinish"}

_CAMPOS_PCT  = {"PercentComplete", "PercentWorkComplete"}

_CAMPOS_WORK     = {"Work"}      
_CAMPOS_DURATION = {"Duration"}   


def _conv_data(dt) -> datetime | None:
    try:
        if dt.year in (1984, 2049):
            return None
        return datetime(dt.year, dt.month, dt.day)
    except Exception:
        return None


def _ler_tarefas(app) -> list[list]:
    atributos = [attr for _, attr in COLUNAS]
    dados = []
    tasks = app.ActiveProject.Tasks

    for i in range(1, tasks.Count + 1):
        try:
            task = tasks.Item(i)
            if task is None:
                continue

            linha = []
            for attr in atributos:
                val = getattr(task, attr, None)

                if attr in _CAMPOS_DATA:
                    val = _conv_data(val)

                elif attr in _CAMPOS_PCT:
                    val = val / 100 if val is not None else None

                elif attr in _CAMPOS_WORK:
                    val = round(val / 60, 2) if val is not None else None

                elif attr in _CAMPOS_DURATION:
                    val = round(val / 480, 2) if val is not None else None

                linha.append(val)

            dados.append(linha)
        except Exception:
            continue

    return dados


def _gravar_xlsx(dados: list[list], caminho: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarefas"

    cabecalhos = [nome for nome, _ in COLUNAS]
    atributos  = [attr for _, attr in COLUNAS]

    ws.append(cabecalhos)

    for linha in dados:
        ws.append(linha)

    for col_idx, attr in enumerate(atributos, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cells = ws[col_letter][1:] 

        if attr in _CAMPOS_DATA:
            for cell in cells:
                cell.number_format = "DD/MM/YYYY"

        elif attr in _CAMPOS_PCT:
            for cell in cells:
                cell.number_format = "0.00%"

        elif attr in _CAMPOS_WORK:
            for cell in cells:
                cell.number_format = '0.00"h"'

        elif attr in _CAMPOS_DURATION:
            for cell in cells:
                cell.number_format = '0.00"d"'

    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if "Nome" in cabecalho:
            ws.column_dimensions[col_letter].width = 50
        else:
            ws.column_dimensions[col_letter].width = max(len(cabecalho) + 2, 12)

    wb.save(caminho)


def exportar_xlsx(app, caminho_mpp: Path) -> Path:
    pasta_xlsx = Path(PASTA_XLSX)
    pasta_xlsx.mkdir(parents=True, exist_ok=True)
    
    stem = caminho_mpp.stem
    if SUFIXO_COPIA:
        if stem.endswith(SUFIXO_COPIA):
            stem = stem[: -len(SUFIXO_COPIA)]

    nome_xlsx = stem + ".xlsx"
    caminho_xlsx = pasta_xlsx / nome_xlsx

    print(f"[EXPORT] Abrindo: {caminho_mpp.name}")
    app.FileOpen(str(caminho_mpp), False)

    try:
        dados = _ler_tarefas(app)
        print(f"  [LIDO] {len(dados)} tarefa(s)")
        _gravar_xlsx(dados, caminho_xlsx)
        print(f"  [XLSX] Salvo: {caminho_xlsx.name}")
    finally:
        app.FileClose(Save=False)

    return caminho_xlsx


def exportar_lista(app, copias: list[Path]) -> list[Path]:
    gerados = []
    for copia in copias:
        try:
            xlsx = exportar_xlsx(app, copia)
            gerados.append(xlsx)
        except Exception as e:
            print(f"[ERRO]   {copia.name}: {e}")
    return gerados


def iniciar_project():
    pythoncom.CoInitialize()
    app = win32com.client.Dispatch("MSProject.Application")
    app.Visible = PROJECT_VISIVEL
    print("[PROJECT] Instância do Microsoft Project iniciado.")
    return app


def encerrar_project(app):
    try:
        app.Quit()
    except Exception:
        pass
    pythoncom.CoUninitialize()
    print("[PROJECT] Instância encerrada.")