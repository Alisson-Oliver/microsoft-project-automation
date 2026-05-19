from pathlib import Path
from datetime import datetime

import win32com.client
import pythoncom
import openpyxl

from config import PROJECT_VISIVEL

COLUNAS = ["ID", "Nome", "Início", "Término", "% Concluído"]


def _ler_tarefas(app) -> list[list]:
    dados = []
    tasks = app.ActiveProject.Tasks
    for i in range(1, tasks.Count + 1):
        try:
            task = tasks.Item(i)
            if task is None:
                continue

            def conv(dt):
                try:
                    return datetime(dt.year, dt.month, dt.day)
                except Exception:
                    return None

            dados.append([
                task.ID,
                task.Name,
                conv(task.Start),
                conv(task.Finish),
                task.PercentComplete,
            ])
        except Exception:
            continue
    return dados


def _gravar_xlsx(dados: list[list], caminho: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarefas"

    ws.append(COLUNAS)

    for linha in dados:
        ws.append(linha)

    fmt_data = "DD/MM/YYYY"
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = fmt_data

    for cell in ws["E"][1:]:
        cell.number_format = "0%"
        if cell.value is not None:
            cell.value = cell.value / 100

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    wb.save(caminho)


def exportar_xlsx(app, caminho_mpp: Path) -> Path:
    caminho_xlsx = caminho_mpp.with_suffix(".xlsx")

    print(f"[EXPORT] Abrindo: {caminho_mpp.name}")
    app.FileOpen(str(caminho_mpp), False)

    try:
        dados = _ler_tarefas(app)
        print(f"  [LIDO] {len(dados)} tarefa(s) encontrada(s)")
        _gravar_xlsx(dados, caminho_xlsx)
        print(f"  [XLSX] Salvo: {caminho_xlsx.name}")
    finally:
        app.FileClose(Save=False)

    return caminho_xlsx


def exportar_lista(app, copias: list[Path]) -> list[Path]:
    gerados = []
    for copia in copias:
        try:
            xlsx = exportar_xlsx(app, copia)
            gerados.append(xlsx)
        except Exception as e:
            print(f"[ERRO]   {copia.name}: {e}")
    return gerados


def iniciar_project():
    pythoncom.CoInitialize()
    app = win32com.client.Dispatch("MSProject.Application")
    app.Visible = PROJECT_VISIVEL
    print("[PROJECT] Instância do Microsoft Project iniciado.")
    return app


def encerrar_project(app):
    try:
        app.Quit()
    except Exception:
        pass
    pythoncom.CoUninitialize()
    print("[PROJECT] Instância encerrada.")


if __name__ == "__main__":
    from buscar_arquivos import listar_mpp
    from copiar_arquivos import copiar_lista

    encontrados = listar_mpp()
    copias = copiar_lista(encontrados)

    app = iniciar_project()
    try:
        gerados = exportar_lista(app, copias)
        print(f"\n[OK] {len(gerados)} arquivo(s) .xlsx gerado(s).")
    finally:
        encerrar_project(app)